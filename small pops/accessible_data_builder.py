from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import pandas as pd

# move files into self.location_of_final_files to run AccessibleData class object

class AccessibleData:
    def __init__(self):
        
        self.location_of_final_files = 'sp-data/accessible-test'
        self.files = [file for file in os.listdir(self.location_of_final_files) if not file.startswith('.')]
        
        self.cantabular_files_path = ""
        self.commission_tables_metadata = ""
        
        self.font_size = 12

        
    def run(self):
        for file in self.files:
            self.accessible_data(file)
            self.accessible_metadata(file)
    
    def accessible_metadata(self, file):
        dataset_id = file.split('.')[0]
        # load data
        book = load_workbook(f"{self.location_of_final_files}/{file}")
        ws = book["Metadata"]
        
        # determine column sizes
        # using preset values
        ws.column_dimensions['A'].width = 44
        ws.column_dimensions['B'].width = 83
        
        # change font size for all
        text_font = Font(size=self.font_size)
        url_font = Font(underline='single', color='000000FF', size=self.font_size)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = text_font
                
        for cell in ws['B']:
            try:
                if cell.value.startswith("=HYPER"):
                    cell.font = url_font
            except:
                pass
        
        # move data 2 rows down
        table_position = ws.dimensions
        ws.move_range(table_position, rows=2)
        
        # add title in cell A1
        dataset_title = self._get_dataset_title(dataset_id)
        ws['A1'] = f"Metadata for {dataset_title}"
        ws['A1'].style = 'Headline 1'
        
        # add value in A2
        ws['A2'] = "This worksheet contains one table of metadata."
        ws['A2'].font = text_font
        
        # table dimensions for table and wrapping text
        table_dimensions = ws.dimensions
        table_dimensions = table_dimensions.replace('A1', 'A3')
        
        # wrap text
        for row in ws[table_dimensions]:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
        
        # create table
        table = Table(displayName="Metadata", ref=table_dimensions)
        table_style = TableStyleInfo(
            name='TableStyleLight1', showFirstColumn=False, 
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
        table.tableStyleInfo = table_style
        ws.add_table(table)
                
        book.save(f"{self.location_of_final_files}/{file}")
        
        return
    
    
    def accessible_data(self, file):
        dataset_id = file.split('.')[0]
        # load data
        book = load_workbook(f"{self.location_of_final_files}/{file}")
        ws = book["Data"]
        
        # determine column sizes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2
        
        # change font size for all
        text_font = Font(size=self.font_size)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = text_font
        
        # move data 2 rows down
        table_position = ws.dimensions
        ws.move_range(table_position, rows=2)
        
        # add title in cell A1
        dataset_title = self._get_dataset_title(dataset_id)
        ws['A1'] = f"Data for {dataset_title}"
        ws['A1'].style = 'Headline 1'
        
        # add value in A2
        ws['A2'] = "This worksheet contains one table."
        ws['A2'].font = text_font
        
        # table dimensions for table and wrapping text
        table_dimensions = ws.dimensions
        table_dimensions = table_dimensions.replace('A1', 'A3')
        
        # wrap text
        for row in ws[table_dimensions]:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
        
        # create table
        table = Table(displayName="Data", ref=table_dimensions)
        table_style = TableStyleInfo(
            name='TableStyleLight1', showFirstColumn=False, 
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
        table.tableStyleInfo = table_style
        ws.add_table(table)
                
        book.save(f"{self.location_of_final_files}/{file}")
        
        return
    
    def _get_dataset_title(self, dataset_id):
        dataset_df = pd.read_csv(f"{self.cantabular_files_path}/Dataset.csv")
        commission_metadata_df = pd.read_excel(f"{self.commission_tables_metadata}", sheet_name="EILR")
        

        if dataset_id.startswith("SP2") and dataset_id.endswith('H') or dataset_id.startswith("SP2") and dataset_id.endswith('G') or dataset_id in ("SP115A", "SP116A", "SP117A", "SP118A", "SP119A"):
            df_loop = commission_metadata_df[commission_metadata_df[' table number'] == dataset_id]
            
            dataset_title = df_loop['table title'].iloc[0]
            
        elif dataset_id.startswith("SP1") or dataset_id.startswith("SP2"):
            df_loop = dataset_df[dataset_df['Dataset_Mnemonic'] == dataset_id]
            
            if len(df_loop) == 0:
                df_loop = dataset_df[dataset_df['Dataset_Mnemonic'] == dataset_id[:-1]]
                if len(df_loop) == 0:
                    raise Exception("df_loop still has length 0..")
            
            dataset_title = df_loop['Dataset_Title'].iloc[0] 
            
        else:
            raise TypeError(f"{dataset_id} should start SP1 or SP2")
            
        return dataset_title

